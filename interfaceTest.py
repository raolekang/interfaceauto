import hashlib
import json
import time
import openpyxl
import requests
import re
import random
import string

class TestInterface:
    param = {
        # os鉴权码
        "AuthorizationCode": "1df5a4b6a5adc2e1f81e95455505626cf419C",
        # 解密 key
        "key": "MIIEvwIBADANBgkqhkiG9w0BAQEFAASCBKkwggSlAgEAAoIBAQDLOElAE93B/ZgDkAJxGvCJb9cu4EaaSsalKOtLKW0OKALP4FsAnu" \
               "CUwj2uUJFxwzlmXA3ML2guPduExl/8hEDgUlFyUUkjV8WrI4LV0iLusz8v4PClk++O20fP8mhowpoKw7G/3cWjBAHxjT6WAzXnHnWLjU8" \
               "H2ejcR6e68xj+ALuxU3Xy73/pPMEn/pjlCRpMBjA3bWoduGVmZiwsO74UD0GpOi2xsYg7fbwUCa+03V4C9hGr2pKtcBcmFUNEOwiAZfXa" \
               "pm+oaDTJJTCtzelOMjqDXITp0BgIGSSrLNZwQWA4tjYtjQtORoq9e2EnIKd8M+RFej+dxX5W+w8p2Z05AgMBAAECggEBAK4QCf3OhlUQp1H1" \
               "2YGTd4/gsck+5HpnrfntMIxycSx87Nus6YWAkBAax8Y00D9G3jVkh5Tqzis9ZHc3iyO5C/CX0Zc/NIwHNBJAGmKW2MDn2ylhW/4W1kvgWoQst2" \
               "1zsYmi0O7sP8o/NoL1XjQTj+nkVv4vDRBGf4OeBHbGnhOAeoTIWGWe3YjyHxFWUnphfPQJoL3MtvM964oGVI2CLhVDnY3H4vaLxNAILpIy96P" \
               "NorUnM2sQx3k1nQDRsjduk1unDTCuwuwz2115De9p9nF4F0PYmXJVLMRUSe8TZaSmMG6h2JeGUuk/s5wtl/Zz1dNTo+WjFA3eXAPJOAW+SgECg" \
               "YEA5fVHul5DxHRoIZlvuISwkpuBUgZsrKCstdvlwZaCKDi9YYuAqHaJ6iv2zZ4JkVFPFRTZb1zPpGvoYZsX6+CuzTZkM12/ScuXGJP3pzJCjH/C6" \
               "B18k/dkeUQxy72DzYYArmcP0vsob/Hpy9SlOsa7xS1XOez0D0gzeRVlUFc9LaECgYEA4jvWL9jA6kaGT14Cd9ibVeQPL1EFK2pTSEE0otzD697ei" \
               "UlhaGpvUh6ug5aL4Yn6FYLiqtFFmpRmGzaNk553TenHMufkkYXYIFC0B4PO+jXrbauNCxoREhgFrHGEWMT8JSaQw9r0YTUAkuRX7rOY2GNgqySC+" \
               "HIEWmspf3RaWJkCgYEAweFeW2RUsJMQpJtwFVtSYBlaHCUshRNm6mF/Qgu6q1zDVzW+AMsIlnYAvPWszWeHTJfqhzjyHIeaQoj3zzj9SCm+q4wz1m0f" \
               "dGp01/PJJhol43vQ6BQiExgK/+NF+AcuFdzw+4okWtXUr7UWpgqej60dqMn+L9BJDaKba/RRsWECgYB9b3Hdi4/DFaFIhjZWHspZwB5PPNvh640Lp" \
               "eO6XS9gK201r3MLwt/AJ3TZTvJ19dusa6qjzW6LO7a1ZfTvoVBC6djht8N6kDx6BmUbOZDcGGNVpmWIf0ZxwtH6f+JCmc24B38CqPcQoqOTznmWxw" \
               "t7BvbFTXQ0rnsNaxdSLMAfGQKBgQCpLYTAoCGKL8XxrOCJAwcTqOmvu4ar7CI9UnG134SV8ndpBm3wh7czLIFHoL1r7agxj/c/6uaS736iiBmtoLrL" \
               "jhBg9vQCAaJkSfJPw2hlCmJIXtFlAT/FrpNJWM1MEXAO+dxKNBYRL0JfKlwmwXzWOImxhbgLKfaz5YH5n4eBog==",
        # 获取一个企业名称
        "companyName": '企业名称序号' + ''.join(random.sample(string.ascii_lowercase + string.digits, 4)),
        # 获取企业简称
        "companySimpleName": "企业简称-" + ''.join(random.sample(string.ascii_lowercase + string.digits, 4)),
        # 获取一个随机账号
        "companyAccount": "cszh" + str(random.randint(1000, 9999)),
        # 获取一个随机chatbotId
        "chatbotId": "afde" + str(random.randint(100000, 999999))
    }

    def getHeader(self,header):
        account = header["account"];
        pwd = header["pwd"];
        timestamp = time.strftime("%m%d%H%M%S", time.localtime())
        str = account.upper() + "00000000" + pwd + timestamp;
        pwd = hashlib.md5(str.encode(encoding='utf-8')).hexdigest();
        header["timestamp"]=timestamp
        header["pwd"]=pwd
        return header

    def getRSA2(self,key,value,type):
        url = "http://218.17.39.34:8888/aim/os/BossService/AgentService/v1/CompanyManage/getRSA2"
        header = {"AuthorizationCode":self.param["AuthorizationCode"] ,"Content-Type":"application/json"}
        body = {
            "key": key,
            "value": value,
            "type": type
        }
        req = requests.request("post",url=url,headers=header,json=body)
        return  re.search(r'"data" : "(.+?)"',req.text)[1]

    def performInterface(self):
        wk = openpyxl.load_workbook('result.xlsx')
        sheet1 = wk["接口测试用例"]
        for x in range(2, sheet1.max_row + 1):
            dict = {}
            for y in range(1, sheet1.max_column + 1):
                dict[sheet1.cell(row=1, column=y).value] = sheet1.cell(row=x, column=y).value
            url = dict["requestUrl"]
            header = json.loads(dict["requestHeader"])
            header = self.getHeader(header)
            body = json.loads(dict["requestBody"])
            if url.endswith("addCompany"):
                 body["companyName"] = TestInterface.param["companyName"]
                 body["companySimpleName"] = TestInterface.param["companySimpleName"]
                 body["mobile"] = TestInterface.getRSA2(self,key=None,value=random.randint(18800000000, 18899999999),type=1)
                 body["companyAccount"] = TestInterface.param["companyAccount"]
            if dict["testResult"] == None or dict["testResult"] == "Failure":
                req = requests.request("post", url=url, headers=header, json=body)
                assertMsg = "Success"
                try:
                    assert req.json()["message"] == dict["checkMsg"]
                except (AssertionError):
                    assertMsg = "Failure"
                finally:
                    if url.endswith("addCompany"):
                        sheet1.cell(row=x, column=6,value = json.dumps(body))
                    sheet1.cell(row=x, column=8,value = req.text)
                    sheet1.cell(row=x, column=9,value = assertMsg)
                    sheet1.cell(row=x, column=10, value= "饶乐康")
                    sheet1.cell(row=x, column=11,value=time.strftime("%Y/%m/%d %H:%M:%S", time.localtime()))
        wk.save("result.xlsx")
        wk.close()


if __name__ == '__main__':
    TestInterface().performInterface()
    print("运行完成")











